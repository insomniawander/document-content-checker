from __future__ import annotations

import argparse
import html
import io
import json
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET


SUPPORTED = {".docx", ".pptx", ".txt", ".md", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"


@dataclass
class Block:
    file: str
    location: str
    text: str


@dataclass
class Difference:
    source_file: str
    source_location: str
    source_text: str
    reference_location: str
    reference_text: str
    similarity: float
    result: str
    important_change: str


def normalize(text: str, ignore_punctuation: bool = True) -> str:
    text = unicodedata.normalize("NFKC", text).replace("\u00a0", " ")
    text = re.sub(r"\s+", "", text)
    if ignore_punctuation:
        text = re.sub(r"[^\w\u4e00-\u9fff]", "", text, flags=re.UNICODE)
    return text.lower()


def meaningful(text: str) -> bool:
    return len(normalize(text)) >= 2


def extract_docx(path: Path) -> list[Block]:
    blocks: list[Block] = []
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    body = root.find("w:body", ns)
    if body is None:
        return blocks
    paragraph_no = table_no = 0
    for child in body:
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "p":
            paragraph_no += 1
            text = "".join((t.text or "") for t in child.findall(".//w:t", ns)).strip()
            if meaningful(text):
                blocks.append(Block(path.name, f"正文第{paragraph_no}段", text))
        elif tag == "tbl":
            table_no += 1
            for row_no, row in enumerate(child.findall(".//w:tr", ns), 1):
                cells = []
                for cell in row.findall("./w:tc", ns):
                    value = "".join((t.text or "") for t in cell.findall(".//w:t", ns)).strip()
                    cells.append(value)
                text = " | ".join(cells).strip(" |")
                if meaningful(text):
                    blocks.append(Block(path.name, f"表格{table_no}第{row_no}行", text))
    return blocks


def extract_pptx(path: Path, enable_ocr: bool = True) -> tuple[list[Block], list[str]]:
    blocks: list[Block] = []
    warnings: list[str] = []
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        slides = sorted(
            (n for n in zf.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)),
            key=lambda n: int(re.search(r"(\d+)", Path(n).stem).group(1)),
        )
        for page, slide_name in enumerate(slides, 1):
            root = ET.fromstring(zf.read(slide_name))
            line_no = 0
            for paragraph in root.findall(".//a:p", ns):
                text = "".join((t.text or "") for t in paragraph.findall(".//a:t", ns)).strip()
                if meaningful(text):
                    line_no += 1
                    blocks.append(Block(path.name, f"第{page}页文本{line_no}", text))
        if enable_ocr:
            media = [n for n in zf.namelist() if n.startswith("ppt/media/") and Path(n).suffix.lower() in IMAGE_EXTS]
            for idx, name in enumerate(media, 1):
                text, error = ocr_bytes(zf.read(name), Path(name).suffix)
                if meaningful(text):
                    for line_no, line in enumerate(split_lines(text), 1):
                        blocks.append(Block(path.name, f"内嵌图片{idx}识别文字{line_no}", line))
                elif error:
                    warnings.append(f"{path.name} / {Path(name).name}: {error}")
    return blocks, warnings


def split_lines(text: str) -> list[str]:
    return [line.strip() for line in re.split(r"[\r\n]+", text) if meaningful(line)]


def find_tesseract() -> str | None:
    found = shutil.which("tesseract")
    candidates = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    return found or next((str(p) for p in candidates if p.exists()), None)


def ocr_bytes(data: bytes, suffix: str) -> tuple[str, str | None]:
    exe = find_tesseract()
    if not exe:
        return "", "未找到 Tesseract OCR，已跳过图片文字"
    with tempfile.TemporaryDirectory(prefix="doc_check_") as temp:
        image_path = Path(temp) / f"image{suffix or '.png'}"
        image_path.write_bytes(data)
        attempts = [[exe, str(image_path), "stdout", "-l", "chi_sim+eng"], [exe, str(image_path), "stdout", "-l", "eng"]]
        last_error = "OCR 执行失败"
        for command in attempts:
            proc = subprocess.run(command, capture_output=True)
            if proc.returncode == 0:
                return proc.stdout.decode("utf-8", errors="replace"), None
            last_error = proc.stderr.decode("utf-8", errors="replace").strip()[-200:]
        return "", last_error


def extract_file(path: Path, enable_ocr: bool = True) -> tuple[list[Block], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return extract_docx(path), []
    if suffix == ".pptx":
        return extract_pptx(path, enable_ocr)
    if suffix in {".txt", ".md"}:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        return [Block(path.name, f"第{i}行", line) for i, line in enumerate(split_lines(text), 1)], []
    if suffix in IMAGE_EXTS:
        text, error = ocr_bytes(path.read_bytes(), suffix) if enable_ocr else ("", "已关闭 OCR")
        blocks = [Block(path.name, f"识别文字第{i}行", line) for i, line in enumerate(split_lines(text), 1)]
        return blocks, ([f"{path.name}: {error}"] if error else [])
    raise ValueError(f"不支持的文件类型: {suffix}")


def important_tokens(text: str) -> set[str]:
    patterns = [
        r"\d+(?:\.\d+)?%",
        r"\d{4}[./年-]\d{1,2}(?:[./月-]\d{1,2}日?)?",
        r"(?:¥|RMB)?\s*\d+(?:\.\d+)?\s*(?:元|万元|亿元|万|亿)?",
    ]
    found: set[str] = set()
    for pattern in patterns:
        found.update(re.findall(pattern, unicodedata.normalize("NFKC", text), flags=re.I))
    return {re.sub(r"\s+", "", x) for x in found if x}


def compare(reference: list[Block], sources: list[Block], threshold: float = 0.88) -> list[Difference]:
    ref_norm = [normalize(b.text) for b in reference]
    results: list[Difference] = []
    for source in sources:
        value = normalize(source.text)
        if not value:
            continue
        best_idx, score = -1, 0.0
        for idx, candidate in enumerate(ref_norm):
            current = 1.0 if value == candidate else SequenceMatcher(None, value, candidate, autojunk=False).ratio()
            if current > score:
                best_idx, score = idx, current
                if score == 1.0:
                    break
        ref = reference[best_idx] if best_idx >= 0 else Block("", "", "")
        src_tokens, ref_tokens = important_tokens(source.text), important_tokens(ref.text)
        token_delta = sorted(src_tokens ^ ref_tokens)
        if score == 1.0:
            result = "一致"
        elif score >= threshold:
            result = "疑似差异"
        elif score >= 0.45:
            result = "不一致"
        else:
            result = "未在基准文档中找到"
        results.append(Difference(
            source.file, source.location, source.text, ref.location, ref.text,
            round(score * 100, 1), result, "、".join(token_delta),
        ))
    return results


def write_xlsx(results: list[Difference], output: Path, warnings: list[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "核对结果"
    headers = ["待核对文件", "位置", "待核对内容", "Word位置", "Word参考内容", "相似度(%)", "结论", "关键数字/日期差异"]
    ws.append(headers)
    for item in results:
        ws.append([item.source_file, item.source_location, item.source_text, item.reference_location,
                   item.reference_text, item.similarity, item.result, item.important_change])
    colors = {"一致": "E2F0D9", "疑似差异": "FFF2CC", "不一致": "FCE4D6", "未在基准文档中找到": "F4CCCC"}
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows(min_row=2):
        fill = PatternFill("solid", fgColor=colors.get(row[6].value, "FFFFFF"))
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [24, 20, 48, 20, 48, 14, 20, 30]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    summary = wb.create_sheet("概要")
    summary.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["核对条数", len(results)])
    for label in ["一致", "疑似差异", "不一致", "未在基准文档中找到"]:
        summary.append([label, sum(r.result == label for r in results)])
    if warnings:
        summary.append(["处理提示", "\n".join(dict.fromkeys(warnings))])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 100
    summary["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(output)


def write_html(results: list[Difference], output: Path, warnings: list[str]) -> None:
    counts = {x: sum(r.result == x for r in results) for x in ["一致", "疑似差异", "不一致", "未在基准文档中找到"]}
    rows = "".join(
        f'<tr class="{html.escape(r.result)}"><td>{html.escape(r.source_file)}</td><td>{html.escape(r.source_location)}</td>'
        f'<td>{html.escape(r.source_text)}</td><td>{html.escape(r.reference_location)}</td><td>{html.escape(r.reference_text)}</td>'
        f'<td>{r.similarity:.1f}%</td><td>{html.escape(r.result)}</td><td>{html.escape(r.important_change)}</td></tr>' for r in results
    )
    warning_html = "".join(f"<li>{html.escape(w)}</li>" for w in dict.fromkeys(warnings))
    page = f"""<!doctype html><meta charset="utf-8"><title>文档内容核对报告</title>
<style>body{{font-family:'Microsoft YaHei',sans-serif;margin:28px;color:#253047}}h1{{color:#17365d}}.cards{{display:flex;gap:12px;flex-wrap:wrap}}.card{{padding:12px 20px;border-radius:10px;background:#edf3f8}}table{{border-collapse:collapse;width:100%;font-size:13px;margin-top:20px}}th,td{{border:1px solid #ccd6df;padding:8px;vertical-align:top}}th{{position:sticky;top:0;background:#1f4e78;color:white}}tr.一致{{background:#e2f0d9}}tr.疑似差异{{background:#fff2cc}}tr.不一致{{background:#fce4d6}}tr.未在基准文档中找到{{background:#f4cccc}}</style>
<h1>文档内容核对报告</h1><p>生成于 {datetime.now():%Y-%m-%d %H:%M:%S}</p>
<div class="cards">{''.join(f'<div class="card"><b>{k}</b><br>{v} 条</div>' for k,v in counts.items())}</div>
{f'<h3>处理提示</h3><ul>{warning_html}</ul>' if warnings else ''}
<table><thead><tr>{''.join(f'<th>{h}</th>' for h in ['待核对文件','位置','待核对内容','Word位置','Word参考内容','相似度','结论','关键差异'])}</tr></thead><tbody>{rows}</tbody></table>"""
    output.write_text(page, encoding="utf-8")


def run_check(reference_path: Path, source_paths: Iterable[Path], output_dir: Path, threshold: float = 0.88, enable_ocr: bool = True) -> tuple[Path, Path, list[Difference], list[str]]:
    reference, warnings = extract_file(reference_path, enable_ocr=False)
    if not reference:
        raise ValueError("基准文档未提取到有效文字")
    sources: list[Block] = []
    for path in source_paths:
        blocks, notes = extract_file(path, enable_ocr=enable_ocr)
        sources.extend(blocks)
        warnings.extend(notes)
    if not sources:
        raise ValueError("待核对文件未提取到文字；如果是图片，请检查 OCR 是否安装")
    results = compare(reference, sources, threshold)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx = output_dir / f"核对报告_{stamp}.xlsx"
    html_path = output_dir / f"核对报告_{stamp}.html"
    write_xlsx(results, xlsx, warnings)
    write_html(results, html_path, warnings)
    return xlsx, html_path, results, warnings


def cli() -> int:
    parser = argparse.ArgumentParser(description="批量核对 PPT/图片与 Word 文档内容")
    parser.add_argument("--reference", required=True, type=Path, help="基准 Word 文档")
    parser.add_argument("--sources", required=True, nargs="+", type=Path, help="待核对文件")
    parser.add_argument("--output", type=Path, default=Path.cwd() / "核对结果")
    parser.add_argument("--threshold", type=float, default=0.88)
    parser.add_argument("--no-ocr", action="store_true")
    parser.add_argument("--json", action="store_true", help="同时在标准输出打印 JSON")
    args = parser.parse_args()
    xlsx, html_path, results, warnings = run_check(args.reference, args.sources, args.output, args.threshold, not args.no_ocr)
    if args.json:
        print(json.dumps({"xlsx": str(xlsx), "html": str(html_path), "results": [asdict(r) for r in results], "warnings": warnings}, ensure_ascii=False))
    else:
        print(f"已生成: {xlsx}\n已生成: {html_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(cli())

